import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from datetime import datetime
from fs_installments.serializers import *
from fs_loans.serializers import *
from fs_transactions.serializers import TransactionSerializer
from fs_users.serializers import UserSerializer
from django.apps import apps

from fs_utils.constants import DATETIME_FORMAT


SERIALIZER_MAPPING = {
    'installment': {'serializer': InstallmentSerializer, 'fields': [
        'ref_number',
        'due_date',
        'principal',
        'interest',
        'total_amount',
        'paid_amount',
        'balance',
        'status',
        'payment_date',
    ]},
    'loan': {'serializer': LoanViewSerializer, 'fields': [
        'ref_number',
        'status',
        'application_number',
        'category_details.name',
        'amount',
        'payment_frequency',
        'loan_term',
        'interest_rate',
        'interest_amount',
        'repayment_type',
        'payment_amount',
        'amount_paid',
        'outstanding_balance',
        'charges',
        'overdue',
        'client_details.display_name',
        'end_date',
        'approved_date',
        'disbursement_date',
        'created_by.display_name',
        'created_at',
    ]},
    'user': {'serializer': UserSerializer, 'fields': [
        'role_name',
        'email',
        'first_name',
        'last_name',
        'phone_number',
        'is_staff',
        'is_active',
        'date_joined',
        'last_login',
    ]},
    'transaction': {'serializer': TransactionSerializer, 'fields': [
        'payment_number',
        'amount',
        'type',
        'ref_number',
        'client_name',
        'created_at',
        'comment',
    ]}
    # Add more mappings as needed
}

MODEL_MAPPING = {
    'user': 'CustomUser'
}


class ExportDataView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        model_name = request.query_params.get('module')

        # Get the model class from the model name
        if not model_name:
            return HttpResponse("Model not found.", status=404)
        app_label = "fs_" + model_name.lower() + "s"
        model = apps.get_model(
            app_label, MODEL_MAPPING.get(model_name) or model_name)

        if not model:
            return HttpResponse("Model not found.", status=404)

        serializer_class = SERIALIZER_MAPPING.get(model_name)
        if not serializer_class:
            return HttpResponse("Serializer not found.", status=404)

        # Fetch data from the database
        queryset = model.objects.all()

        # Serialize the data
        serializer = serializer_class['serializer'](queryset, many=True)
        data = serializer.data

        # Create a new Excel workbook and add a sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = model_name

        # Get model fields to use as headers
        # fields = serializer_class.Meta.fields
        fields = serializer_class['fields']
        worksheet.append(fields)

        # Populate the worksheet with data
        for item in data:
            row = []
            for field in fields:
                if '.' in field:
                    # Handle nested fields
                    parts = field.split('.')
                    value = item
                    for part in parts:
                        value = value.get(part, {})
                    row.append(value if value != {} else '')
                else:
                    # Handle non-nested fields
                    row.append(item.get(field, ''))
            worksheet.append(row)

        # Auto-size columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Generate filename
        filename = f'{model_name}s-export-{datetime.today().strftime(DATETIME_FORMAT)}.xlsx'

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # Save workbook to response
        workbook.save(response)
        return response
